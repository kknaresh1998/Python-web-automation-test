from traceback import print_tb

import openpyxl
from selenium import  webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.firefox.service import Service as FirefoxService

def Excel_File_Update(file_path, fruit_Name, column_Name, newValue ):
    # Edit the price of fruit in excel file and upload
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1,column=i).value == column_Name:
            Dict["col"] = i

    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == fruit_Name:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = newValue
    book.save(file_path)

driver = webdriver.Firefox(service=FirefoxService(GeckoDriverManager().install()))

fruit_Name = "Apple"
column_Name = "Price"
newValue = 500
file_Path = r"C:\Users\admin\Downloads\download(1).xlsx"
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()
driver.implicitly_wait(5)

driver.find_element(By.ID,"downloadButton").click()

Excel_File_Update(file_Path, fruit_Name, column_Name, newValue)

file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_Path)

toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait = WebDriverWait(driver, 5)
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

price_Column = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_Price = driver.find_element(By.XPATH,"//div[text()='"+fruit_Name+"']/parent::div/parent::div/div[@id='cell-"+price_Column+"-undefined']")
print(actual_Price.text)
assert actual_Price == newValue


