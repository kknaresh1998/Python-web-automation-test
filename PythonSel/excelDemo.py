from email.header import decode_header
from traceback import format_tb
Dics ={}
import openpyxl
book = openpyxl.load_workbook("C:\\Users\\admin\\Documents\\ExcelDemoPython.xlsx")
sheet = book.active
cell= sheet.cell(row=1, column=1)
print(cell.value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "testcase 2":
        for j in range(1, sheet.max_column+1):
            #print(sheet.cell(row=i,column=j).value)
            Dics[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
print(Dics)


